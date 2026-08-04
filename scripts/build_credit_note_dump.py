"""Converts a raw Zoho Books credit note export (.xlsx, sheet "CreditNotes")
into frontend/data/creditnotedump.json — the exact columns Zoho gives us,
unchanged, for the Credit Notes tab's read-only raw view.

Mirrors scripts/build_invoice_dump.py: deliberately separate from the
existing creditnotes.json that still drives Net revenue on Recurring
Revenue/product tabs (Gross - Credit) — that dataset is untouched. This one
is just a faithful mirror of what Zoho actually exported.

Run:  py -3 scripts/build_credit_note_dump.py "<path to CreditNotes export.xlsx>"
"""
import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

MN = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

COLS = [
    ("cndate", "Credit Note Date"), ("cn", "Credit Note Number"), ("status", "Credit Note Status"),
    ("client", "Customer Name"), ("total", "Total"), ("einvoice", "e-Invoice Status"),
    ("assocInv", "Associated Invoice Number"), ("assocInvDate", "Associated Invoice Date"),
    ("currency", "Currency Code"), ("fx", "Exchange Rate"), ("item", "Item Name"), ("desc", "Item Desc"),
    ("qty", "Quantity"), ("price", "Item Price"), ("itemTotal", "Item Total"), ("taxable", "Taxable Amount"),
    ("usageFrom", "Item.CF.Usage Period (From)"), ("usageTill", "Item.CF.Usage Period (till)"),
    ("supplierGstin", "Supplier GST Registration Number"), ("taxPct", "Item Tax %"), ("hsn", "HSN/SAC"),
    ("cgst", "CGST"), ("sgst", "SGST"), ("igst", "IGST"), ("branch", "Branch Name"),
    ("gstTreatment", "GST Treatment"), ("gstin", "GST Identification Number (GSTIN)"),
    ("discount", "Discount Amount"), ("ref", "Reference#"),
]
NUMERIC = {"total", "fx", "qty", "price", "itemTotal", "taxable", "taxPct", "cgst", "sgst", "igst", "discount"}
DATE_FIELDS = {"cndate", "assocInvDate", "usageFrom", "usageTill"}


def fmt_date(v):
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return f"{v.day:02d}-{MN[v.month - 1]}-{str(v.year)[2:]}"
    return str(v)


def fmt_num(v):
    if v is None or v == "":
        return 0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0


def main():
    if len(sys.argv) < 2:
        print("usage: py -3 scripts/build_credit_note_dump.py <path to CreditNotes export.xlsx>")
        sys.exit(1)
    src = Path(sys.argv[1])
    out = Path(__file__).resolve().parent.parent / "frontend" / "data" / "creditnotedump.json"

    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {label: header.index(label) for _, label in COLS if label in header}
    missing = [label for _, label in COLS if label not in header]
    if missing:
        print("WARNING - columns not found in source file:", missing)

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        row = {}
        for key, label in COLS:
            if label not in idx:
                row[key] = "" if key not in NUMERIC else 0
                continue
            v = r[idx[label]]
            if key in DATE_FIELDS:
                row[key] = fmt_date(v)
            elif key in NUMERIC:
                row[key] = fmt_num(v)
            else:
                row[key] = "" if v is None else str(v)
        rows.append(row)

    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps({
        "cols": [k for k, _ in COLS],
        "labels": {k: label for k, label in COLS},
        "rows": rows,
    }), encoding="utf-8")
    print(f"wrote {len(rows)} rows, {len(COLS)} columns -> {out}")


if __name__ == "__main__":
    main()
